import os.path as osp


class TableRenderer(object):
    """
    creating a prettier table
    """

    def __init__(self, col_headers, name="",
                 first_col_width=25, col_width=15,
                 as_md=False, verbose=False):
        self.col_headers = col_headers
        self.name = name
        self.rows = []
        self.first_col_width = first_col_width
        self.col_width = col_width
        self.as_md = as_md
        self.verbose = verbose

    def add_row(self, row_name, content):
        if isinstance(content, (list, tuple)):
            assert len(content) == len(self.col_headers), \
                "content columns mismatch"
            row_content = content
        elif isinstance(content, dict):
            row_content = []
            for col_name in self.col_headers:
                if self.verbose and col_name not in content:
                    print(f"{col_name} is not in content {content}")
                value = content.get(col_name, "N.A.")
                row_content.append(value)
        else:
            raise ValueError("invalid content")
        self.rows.append([row_name] + row_content)

    def __str__(self):
        template = ["{:<{w0}}"]
        template += ["{:^{w1}}" for _ in self.col_headers]
        template = " | ".join(template)
        if not self.as_md:
            lines = [self.name]
            col_headers = [""] + self.col_headers
            headers = [col_headers]
            first_col_width = self.first_col_width
        else:
            lines = []
            first_col_width = max(self.first_col_width, len(self.name)+2)
            col_headers = [self.name] + self.col_headers
            splitter = [":"+"-"*(first_col_width-1)] + \
                ["-"*self.col_width]*len(self.col_headers)
            headers = [col_headers, splitter]
        for row in headers + self.rows:
            row_str = template.format(*row,
                                      w0=first_col_width,
                                      w1=self.col_width)
            lines.append(row_str)
        return "\n".join(lines)

    def to_excel(self, save_path):
        try:
            from openpyxl import Workbook
            from openpyxl import load_workbook
        except ImportError:
            print("needs openpyxl, please pip install")
            return
        if not save_path.endswith(".xlsx"):
            save_path = save_path + ".xlsx"
        if osp.exists(save_path):
            wb = load_workbook(filename=save_path)
        else:
            wb = Workbook()
        if self.name in wb:
            print(f"[Warning] {save_path} already has worksheet {self.name}")
        ws = wb.create_sheet(title=self.name)
        col_headers = [""] + self.col_headers
        for row in [col_headers] + self.rows:
            ws.append(list(map(str, row)))
        wb.save(filename=save_path)
        print(f"save '{self.name}' to {save_path}")
